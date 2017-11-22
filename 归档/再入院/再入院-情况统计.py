# -*- coding: utf-8 -*-

'''
各科室1日内，2-31日内再入院数量统计
依靠已有模板

'''

import pandas as pd
from openpyxl.reader.excel import load_workbook
import numpy as np


text = "请将[全院.xls]和[全院再入院汇总.xls]另存为.xlsx格式的文件\n修改[模板.xlsx],并确保以上文件与脚本位于同目录下\n输入y继续"
x = ''
while not x == 'y':
    x = input(text


df = pd.read_excel(r'全院.xlsx')
df_sum = load_workbook(filename = r'全院再入院汇总.xlsx')
wb_sum = df_sum.get_active_sheet()
wb = load_workbook(filename=r'模板.xlsx',data_only = True)
ws = wb.get_sheet_by_name('质控简报用表-例数统计')
compare_ws = wb.get_sheet_by_name('科室名称对照')

ks_dict = {}
for row in compare_ws.rows:
    ks_dict[row[1].value]=row[2].value



firstin = df[df['同病再入院'].isnull()]
not_firstin = df[df['同病再入院'].notnull()]
not_firstin = not_firstin[not_firstin.同病再入院 < 5]

grouped_firstin = firstin.groupby(['出院科别'])
grouped_zry = grouped_firstin.count()['病历号']

nrow1day = 6
nrow31day = 6
new_row = 0 
#遍历trans_wb中科室的名字
for nrow in compare_ws.rows:
    new_row = max(nrow1day,nrow31day,new_row)
    out_name = nrow[0].value
    start = new_row
    if out_name == None:
        continue

    try:
        name = ks_dict[out_name]
        ws['A{}'.format(new_row)] = out_name
        
        #获得统计表出院数
        for row in wb_sum.rows:
            if row[0].value == name:
                zj = row[1].value

        
        ws['B{}'.format(new_row)] = zj
        
        #计算全部再入院数
        try:
            ws['C{}'.format(new_row)] = grouped_zry[name]

#无再入院科室的处理
        except KeyError:
            end = new_row
            ws['A{}'.format(new_row)].value= None
            ws['B{}'.format(new_row)].value= None
            #ws['C{}'.format(new_row)] = 0
            #new_row +=1
         ###   print('start:{} end:{} name:{}'.format(start,end,out_name))
            continue
            
        nrow1day = new_row
        nrow31day = new_row
        
        #是否有该科室的再入院数据
        try: 
            
            firstin_depart = grouped_firstin.get_group(name)
            #科室首次再入院患者病历号
            firstin_depart_iid = firstin_depart['病历号']
            #遍历病历号
            count1 = 0 #计数用
            count31 = 0
            
            #遍历该科室的病人
            for n in range(len(firstin_depart_iid)):
                iid = firstin_depart_iid.iloc[n]
                #获得该病历号信息
                info = not_firstin[not_firstin.病历号 == iid]
                if info.size == 0: #无该病人数据
                    continue
                
                #判断住院间隔，写入
                
                if info['住院间隔'].iloc[0] < 2 :
                    ws['E{}'.format(nrow1day)].value = iid
                    ws['F{}'.format(nrow1day)].value = info['同病再入院'].iloc[0]
                    nrow1day += 1
                    count1 +=1
                elif info['住院间隔'].iloc[0] < 32:
                    ws['H{}'.format(nrow31day)].value = iid
                    ws['I{}'.format(nrow31day)].value = info['同病再入院'].iloc[0]
                    nrow31day +=1
                    count31 +=1
#无4类出院患者处理
            if count1 == count31 ==0:
                ws['A{}'.format(new_row)].value= None
                ws['B{}'.format(new_row)].value= None
                ws['C{}'.format(new_row)].value= None
                continue
                #new_row +=1
                    
            if count1 > 0:
                ws['D{}'.format(new_row)] = count1

            if count31 > 0:
                ws['G{}'.format(new_row)] = count31

                
         
            end = max(nrow1day,nrow31day,new_row) - 1
            
            ###print('start:{} end:{} name:{}'.format(start,end,out_name))

            if end > start:
                ws.merge_cells("A{}:A{}".format(str(start),str(end)))
                ws.merge_cells("B{}:B{}".format(str(start),str(end)))
                ws.merge_cells("C{}:C{}".format(str(start),str(end)))
            
        except KeyError:
            new_row += 1
            continue
        
        
        
            
        
    except KeyError:
        continue

#全院数据与计数
ws['A5'].value = '全院'
ws['B5'].value = wb_sum['B2'].value

all_rein = 0
all_1day = 0
all_31day = 0
for n in ws.rows:
    print('D:',n[3].value,type(n[3].value))
    if type(n[2].value) == np.int64:
        all_rein += n[2].value
    if type(n[3].value) == int:
        all_1day += n[3].value
    if type(n[6].value) == int:
        all_31day +=n[6].value


ws['C5'].value = all_rein
ws['D5'].value = all_1day
ws['G5'].value = all_31day
    

 
wb.remove_sheet(compare_ws)
wb.save(r'情况统计.xlsx')