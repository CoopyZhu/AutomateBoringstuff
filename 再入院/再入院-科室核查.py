# -*- coding: utf-8 -*-
"""
需要科室解释的再入院情况统计

@author: Coopy
"""
import pandas as pd
from pandas import DataFrame
import openpyxl
from openpyxl.reader.excel import load_workbook


df = pd.read_excel(r'全院.xlsx')
wb = openpyxl.Workbook()
ws = wb.active
ws.tittle='需要科室解释的再入院情况统计'

col = ['病历号','首次出院科室','首次出院病房','首次入院时间','首次入院诊断'
       ,'首次出院时间','再入院病房','再入院时间','再入院诊断','住院次数判断','类别','住院间隔']

for n in range(1,13):
    ws.cell(row=1,column=n).value = col[n-1]

nrow = 2
for nset in range(1,5):
    set_n = df[df.同病再入院 == nset]
    print(nset,set_n.size)
    for iid in set_n['病历号'].unique():
        set_id = df[df.病历号 == iid]
        
        r0 = set_id.iloc[0]
        r1 = set_id.iloc[1]
        ws.cell(row=nrow,column=1).value = str(iid)
        ws.cell(row=nrow,column=2).value = r0['出院科别']
        ws.cell(row=nrow,column=3).value = r0['出院病房']
        ws.cell(row=nrow,column=4).value = r0['入院日期']
        ws.cell(row=nrow,column=4).number_format='yyyy/m/d'
        ws.cell(row=nrow,column=5).value = r0['s050100']
        ws.cell(row=nrow,column=6).value = r0['出院日期']
        ws.cell(row=nrow,column=6).number_format='yyyy/m/d'
        ws.cell(row=nrow,column=7).value = r1['入院病房']
        ws.cell(row=nrow,column=8).value = r1['入院日期']
        ws.cell(row=nrow,column=8).number_format='yyyy/m/d'
        ws.cell(row=nrow,column=9).value = r1['s050100']
        ws.cell(row=nrow,column=10).value = len(set_id['病历号'])
        ws.cell(row=nrow,column=11).value = r1['同病再入院']
        ws.cell(row=nrow,column=12).value = r1['住院间隔']

        nrow +=1 
    
wb.save(r'需要科室解释原因的再入院判断.xlsx')



