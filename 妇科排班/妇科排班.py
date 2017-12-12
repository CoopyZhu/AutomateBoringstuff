# -*- coding: utf-8 -*-
"""
Created on Mon Dec 11 16:43:30 2017
妇科排班

@author: Coopy
"""
from openpyxl import load_workbook


def getstaff():

    ws = wb.get_sheet_by_name('人员')
    name_dict ={}
    for row in ws.iter_rows():
        level = row[0].value
        name_list = []
        for cell in row[1:]:
            if cell.value != None:
                name_list.append(cell.value)
        name_dict[level] = name_list
    return name_dict

#以列表形式接受某组一周的排班信息,存为星期对应的字典
def getoneweek(aweek):
    oneweek = {}
    for d in range(1,8):
        oneweek[d] = aweek[d-1]

    return oneweek

#接受小组名，获得其前三周夜班信息,返回白、夜、行索引
def get3weeks(xz):
    ws = wb.get_sheet_by_name('排班')
    weeks_day = []
    weeks_night = []
    for row in ws.iter_rows():
        if xz in str(row[0].value):
            row_idx = row[0].row
            if '夜' in str(row[0].value):
                row_idx_night = row[0].row
                for col in ws.iter_cols(1):
                    weeks_night.append(col[row_idx-1].value)
            if '白' in str(row[0].value):
                 row_idx_day = row[0].row
                 for col in ws.iter_cols(1):
                    weeks_day.append(col[row_idx-1].value)
    weeks_day.pop(0)
    weeks_night.pop(0)
    return weeks_day,weeks_night,row_idx_day,row_idx_night

def bjget3weeks(xz):
    ws = wb.get_sheet_by_name('排班')
    weeks_night = []
    for row in ws.iter_rows():
        if xz in str(row[0].value):
            row_idx = row[0].row
            if '夜' in str(row[0].value):
                row_idx_night = row[0].row
                for col in ws.iter_cols(1):
                    weeks_night.append(col[row_idx-1].value)

    weeks_night.pop(0)
    return weeks_night,row_idx_night

def scheduleone(xz,num):
    weeks_day,weeks_night,row_idx_day,row_idx_night = get3weeks(xz)                   
    week1_day = getoneweek(weeks_day[0:7])
    week1_night = getoneweek(weeks_night[0:7])
    week2_day = getoneweek(weeks_day[7:14])
    week2_night = getoneweek(weeks_night[7:14])
    week3_day = getoneweek(weeks_day[14:21])
    week3_night = getoneweek(weeks_night[14:21])
    
    #初始化新一周
    week0_night = {}
    week0_day = {}
    
    generalrule(week0_night,week3_night)
    
    if num == 6:
        week0_day,week0_night = rule6s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night)
    elif num == 7:
        week0_day,week0_night = rule7s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night)
    elif num == 8:
        week0_day,week0_night = rule8s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night)
    elif num ==9:
        week0_day,week0_night = rule9s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night)
    elif num == 10:
        week0_day,week0_night = rule10s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night)
    elif num == 11:
        week0_day,week0_night = rule11s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night)
    elif num == 12:
        week0_day,week0_night = rule12s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night)
        
    #写入表格
    ws = wb.get_sheet_by_name('排班')
    for col_idx in range(23,30):
        d = col_idx - 22
        try:
            ws.cell(row=row_idx_day,column = col_idx).value = week0_day[d]
        except KeyError:
            pass
        ws.cell(row=row_idx_night,column = col_idx).value = week0_night[d]

def schedulebj(xz,num):
    weeks_night,row_idx_night = bjget3weeks(xz)
    week1_night = getoneweek(weeks_night[0:7])
    week2_night = getoneweek(weeks_night[7:14])
    week3_night = getoneweek(weeks_night[14:21])
    week0_night = {}
    generalrule(week0_night,week3_night)
    if num == 7:
        week0_night[7] = week3_night[1]
    elif num == 8:
        week0_night[7] = week2_night[1]
    elif num ==9 :
        week0_night[7] = week1_night[1]
    else:
        week0_night[7] = None
    ws = wb.get_sheet_by_name('排班')
    for col_idx in range(23,30):
        d = col_idx - 22

        ws.cell(row=row_idx_night,column = col_idx).value = week0_night[d]
     
def scheduleall():
    name_dict = getstaff()
    for key in name_dict.keys():
        num = len(name_dict[key])
        if '备叫' in key:
            schedulebj(key,num)
        else:
            scheduleone(key,num)
    return

def generalrule(week0_night,week3_night):

    for key in week3_night.keys():
        week0_night[key-1] = week3_night[key]
    week0_night.pop(0)
    return week0_night

def rule6s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night):
    week0_night[7] = week0_night[1]
    week0_day[6] = week0_night[2]
    week0_day[7] = week0_night[4]
    return week0_day,week0_night
            
def rule7s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night):
    week0_night[7] = week0_night[1]
    week0_day[6] = week0_night[3]
    week0_day[7] = week3_night[5]
    week0_night[4] = week3_day[7]
    return week0_day,week0_night

def rule8s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night):
    week0_night[7] = week0_night[1]
    week0_day[6] = week3_night[3]
    week0_day[7] = week3_night[5]
    week0_night[2] = week3_day[6]
    week0_night[4] = week3_day[7]
    return week0_day,week0_night

def rule9s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night):
    week0_night[7] = week0_night[1]
    week0_day[6] = week3_night[3]
    week0_day[7] = week3_night[5]
    week0_night[2] = week2_day[6]
    week0_night[4] = week3_day[7]
    return week0_day,week0_night

def rule10s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night):
    week0_night[7] = week3_night[1]
    week0_day[6] = week3_night[3]
    week0_day[7] = week3_night[5]
    week0_night[2] = week2_day[6]
    week0_night[4] = week3_day[7]
    return week0_day,week0_night

def rule11s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night):
    week0_night[7] = week3_night[1]
    week0_day[6] = week3_night[3]
    week0_day[7] = week3_night[5]
    week0_night[2] = week2_day[6]
    week0_night[4] = week2_day[7]
    return week0_day,week0_night

def rule12s(week0_day,week0_night,week1_day,week1_night,week2_day,week2_night,week3_day,week3_night):
    week0_night[7] = week3_night[1]
    week0_day[6] = week3_night[3]
    week0_day[7] = week3_night[5]
    week0_night[2] = week2_day[6]
    week0_night[4] = week2_day[7]
    week0_night[3] = week2_night[4]
    return week0_day,week0_night

input('''请将【排班.xlsx】文件放于脚本同目录下，
该脚本使用前三周的排班表作为依据，每运行一次，生成新一周排班表
因此，每生成一周排班后，请删除最前一周排班，方可再次运行该脚本
继续请输入任意键''')
wb = load_workbook('排班.xlsx')
scheduleall()
wb.save('排班.xlsx')
print('已保存至【排班.xlsx】')
input('输入任意键退出')