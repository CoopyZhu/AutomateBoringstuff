# -*- coding: utf-8 -*-
"""
Created on Mon Oct  8 15:15:44 2018

@author: 朱诚锐

提取信息
"""
import jieba
import jieba.posseg as pseg
from openpyxl import load_workbook
import re
#载入自定义词典
jieba.load_userdict("User_Dict.txt")
#文件路径
path = r"C:\Users\朱诚锐\OneDrive\DEEPWISE\问诊平台\数据\征象提取"
#文件名
filename = r"\z-少突星形细胞瘤20LL.xlsx"
dirc = path+filename

book = load_workbook(dirc)
sheet = book["Sheet1"]


    
#提取信号
'''
def WriteTo(w,row,result):
    if w.flag=="xhmc":
        xhmc = ["T1WI","T2WI","FLAIR","DWI","","","","ADC"]
        if w.word in xhmc:
            for ww in result:
                if ww.flag == "xhqd":
                    ww.word = ww.word +"信号"
                    existed_cell = sheet.cell(row=row,column=13+xhmc.index(w.word))
                    if existed_cell.value is None:
                        existed_cell.value=ww.word
                    else:
                        if ww.word in existed_cell.value:
                            next
                        else:
                            existed_cell.value =  existed_cell.value +"、" +ww.word
        

                            
#遍历每行内容                         
for row in range(2,sheet.max_row+1):
    text = sheet.cell(row=row,column=9).value
    if not text is None:
        #根据，。分隔为list
        describe = re.split(r"，|。|,",text)
        #遍历每句话
        for sentence in describe:
            print("sentence:",sentence)
            result = pseg.cut(sentence,HMM=False)
            #遍历该句中每个词
            for w in result:
                print("word:{} flag:{}".format(w.word,w.flag))
                #如果该句中有信号名称
                if w.flag =="xhmc":
                    #若该句中有顿号
                    sentences = sentence.split("、")
                    if "、"in sentence:
                        #遍历切分顿号后的每句
                        for sentence2 in sentences:
                            result2 = pseg.cut(sentence2,HMM=False)
                            for w2 in result2:
                                WriteTo(w2,row,result2)
                                continue
                            continue
                        break
                    else:
                        WriteTo(w,row,result)
                if w.flag =="xhxz":
                    existed_cell = sheet.cell(row=row,column=25)
                    if existed_cell.value is None:
                        existed_cell.value=w.word
                    else:
                        if w.word in existed_cell.value:
                            next
                        else:
                            existed_cell.value =  existed_cell.value +"、" +w.word
                
'''

#保存
book.save(path+"\z-少突星形细胞瘤20LL处理后.xlsx")
'''

'''