# -*- coding: utf-8 -*-
"""
Created on Fri Oct 26 19:57:54 2018

@author: 朱诚锐
"""

from openpyxl import load_workbook
import re


file_dir = r"C:/Users/朱诚锐/OneDrive/DEEPWISE/问诊平台/数据/"
file_name ="详情补充.xlsx"

file_path = file_dir + file_name

wkb = load_workbook(file_path)
sheet = wkb.active

for row in range(2,sheet.max_row+1):
    content = sheet.cell(row=row,column=2).value
    content = content.strip("\n")
    content = re.sub(r"\n\n+","\n",content)
    content = content.split("\n")
    now_col = 3
    con=""
    for c in content:
        if len(c)<10:
            sheet.cell(row=row,column=now_col).value = con
            con = ""
            
            now_col +=1
            con = con + c +"\n"
            now_col +=1
        else:
            con= con +c+"\n"
    sheet.cell(row=row,column=now_col).value = con

wkb.save("详情补充1.xlsx")


