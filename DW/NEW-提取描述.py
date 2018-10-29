# -*- coding: utf-8 -*-
"""
Created on Wed Oct 10 17:04:54 2018

@author: 朱诚锐
"""
import jieba
import jieba.posseg as pseg
import re
from openpyxl import load_workbook

jieba.load_userdict(r"User_Dict.txt")


def Cut_byAnatomy(Content):
    '''
    将描述原文根据解剖位置切分成段落
    判断
    一个set为（、）（并列关系）（位置词）解剖部位     
    识别解剖部位，若其前面不是顿号、位置词或解剖部位，则进行分段
    '''
    if not isinstance(Content,str):
        raise TypeError("not string type")
    Content = Content.replace(" ","")
    outp = []#由切片组成，短语的slic表示
    slic = ""#由jieba分词结果组成
    
    result = pseg.cut(Content, HMM=False)
    r_list = []
    #将分词结果存为list[[word,flag],]
    for w in result:
        r_list.append([w.word,w.flag])
    
    for i in range(len(r_list)):
        w = r_list[i]
        w_before = r_list[i-1]
        try:
            w_after = r_list[i+1]
        except:
            w_after = ["",""]
        if i==0:
            w_before=["",""]
        #如果该词是 解剖 或 位置，对其前置词进行分析
        if w[1] in ["jp"]:
            #若其前面不是顿号、位置词或解剖部位，则将slic移入outp，从该词新建slic
            if (w_before[1] not in ["jp","wz","blgx"]) and w_before[0] not in["、","-"]:
                outp.append(slic)
                slic = w[0]
            else:
                slic += w[0]
        elif w[1] in ["wz"]:
            if w_after[1] not in ['jp',"wz","blgx"]:
                slic += w[0]
            else:
                if (w_before[1] not in ["jp","wz","blgx"]) and w_before[0] not in["、","-"]:
                    outp.append(slic)
                    slic=w[0]
                else:
                    slic +=w[0]               
        else:
            slic+= w[0]
    #End Slic
    outp.append(slic)
    while "" in outp:
        outp.remove("")
            
        return outp

def Cut_byXHMC(Text):
    '''
    处理的结果，提取其中的信号描述
    '''
    result = pseg.cut(Text,HMM=False)
    #将flag和word分别存为list
    flags = []
    words = []
    for w in result:
        flags.append(w.flag)
        words.append(w.word)
    if "xhmc" in flags:
        pass
    
def WriteToCell(Cell,Value,quto="、",override=False):
    """
    将内容写入单元格，去重，中间用quto分隔
    """
    if override == True:
        Cell.value = Value
    else:
        if Cell.value ==None or Cell.value =="":
            Cell.value = Value
        else:
            if Value not in Cell.value:
                Cell.value += (quto+Value)
            
class Word:
    def  __init__(self,text,flag,index):
        self.word = text
        self.flag = flag
        self.index = index
class Sentence:
    def __init__(self,index):
        self.index = index
        self.result = []
        self.words = []
        self.flags = []
    def add(self,each_Word):
        self.result.append(each_Word)
        self.words.append(each_Word.word)
        self.flags.append(each_Word.flag)

    def before(self,a_Word):
        idx = a_Word.idnex
        if idx ==0:
            return None
        else:
            return Sentence(self.index,self.result[:idx],self.words[:idx],self.flags[:idx])
            
class ParsedContent:
    def __init__(self,raw_content):
        # 将内容按照标点符号分隔
        if raw_content == None:
            self.contents = None
        else:
            self.contents=re.split(r",|，|。|、|;|；|/",raw_content.strip())
            self.sentences=[]
            for content in self.contents:
                result = pseg.cut(content,HMM=False)
                s_idx = 0
                w_idx = 0
                sentence = Sentence(index=s_idx)
                for w in result:                    
                    word=Word(w.word,w.flag,w_idx)
                    sentence.add(word)
                    w_idx += 1
                self.sentences.append(sentence)
                s_idx +=1      
    
class ParsedDiagnosis:
    def __init__(self,raw_diagnosis):
        if raw_diagnosis ==None:
            self.diagnosis = [""]
        else:
            self.diagnosis = re.split(r"\n",raw_diagnosis.strip())
    def if_occupied(self,disease,write_to_cell,key_word = "占位"):
        for each_dig in self.diagnosis:
            if disease in each_dig:
                if key_word in each_dig:
                    write_to_cell.value = "有"
    def parse_jp(self, disease,write_to_cell):
    #处理诊断中的解剖位置
        for each_dig in self.diagnosis:
            if disease in each_dig:
                result = pseg.cut(each_dig,HMM=False)
                for w in result:
                        if w.flag == "jp":
                            WriteToCell(write_to_cell,w.word,"、")
                            
                            
'''
TODO：单位句子中存在两个及以上信号名称如何处理，难点-长T1，T1低信号，顺序问题
'''

    
def main():
    path = r"C:\Users\朱诚锐\OneDrive\DEEPWISE\问诊平台\数据\征象提取"
    #文件名
    filename = r"\1022-颅内海绵状血管瘤.xlsx"
    #输出文件名
    out_filename = r"\1022处理后.xlsx"
    dirc = path+filename

    book = load_workbook(dirc)
    sheet = book["Sheet1"]
    
    #初始化信号位置
    content_col=13 #空白列，作为存储提取描述内容的容器
    raw_content_col = 5 #原始诊断内容
    diagnosis_col = 6
    jp_col = 14
    xhxz_col = 28
    zqsm_col = 20
    zqsmlx_col = 21
    zw_col = 31
    disease="海绵" #疾病关键字
    T1_col = 16
    T2_col=17
    Flair_col=18
    DWI_col = 19
    SWI_col = 24
    df_col = 15
    ADC_col = 23
    
    #遍历数据行,
    for row in range(2,sheet.max_row+1):
        #加载目标单元格
        diagnosis_cell = sheet.cell(row=row,column=diagnosis_col)
        content_cell = sheet.cell(row=row,column=content_col)
        raw_content = sheet.cell(row=row,column=raw_content_col).value
        xhxz_cell = sheet.cell(row=row,column=xhxz_col)
        zqsm_cell = sheet.cell(row=row,column=zqsm_col)
        zqsmlx_cell = sheet.cell(row=row,column=zqsmlx_col)
        jp_cell = sheet.cell(row=row,column=jp_col)
        zw_cell = sheet.cell(row=row,column=zw_col)
        T1_cell=sheet.cell(row=row,column=T1_col)
        T2_cell=sheet.cell(row=row,column=T2_col)
        Flair_cell=sheet.cell(row=row,column=Flair_col)
        SWI_cell=sheet.cell(row=row,column=SWI_col)
        DWI_cell=sheet.cell(row=row,column=DWI_col)
        df_cell = sheet.cell(row=row,column=df_col)
        ADC_cell = sheet.cell(row=row,column=ADC_col)
        print("row:{}\nraw_content:{}\n{}".format(row,raw_content,"+"*30))
        
        
        #处理诊断，
        diagnosis = ParsedDiagnosis(diagnosis_cell.value)
        #识别疾病所在的诊断内容，提取解剖位置，写入jp_cell
        diagnosis.parse_jp(disease,jp_cell)
        #识别疾病所在的诊断内容，提取是否占位，写入zw_cell
        diagnosis.if_occupied(disease,zw_cell)
        #识别疾病所在的诊断内容，提取是否多发，写入df_cell
        diagnosis.if_occupied(disease,df_cell,"多发")
        
        
        #处理描述，将疾病部位所在的段落提取出
        try:
            jp = jp_cell.value.split("、")
            print("jp:",jp)
            for each_jp in jp:
                for sentence in (Cut_byAnatomy(raw_content)):
                    if each_jp in sentence:
                        WriteToCell(content_cell,sentence,quto="。")
        except:
            content_cell.value = raw_content
        

        content = sheet.cell(row=row,column=content_col).value
        print("content:",content)

        #根据标点符号拆分内容
        class_content = ParsedContent(content)
        if class_content.contents == None:
            continue
        else:
            for sentence in class_content.sentences:
                result = sentence.result
                flags = sentence.flags
                words = sentence.words
                if sentence ==None or words == None:
                    continue
                print(words)
                try:
                    #对增强扫描进行处理
                    if "zqsm" in flags:
                        #print("flags:{} \nwords:{}".format(flags,words))
                        #如果存在zqsm，同时有存在性负的提示词，则记为不强化
                        if "czxf" in flags and "xhmc" not in flags:
                            WriteToCell(zqsm_cell,"不强化")
                        elif "czxz" in flags and"czxf" not in flags:
                            WriteToCell(zqsm_cell,words[flags.index("czxz")]+"强化")
                            if "xhxr" in flags:
                                WriteToCell(zqsmlx_cell,words[flags.index("xhxr")]+"强化")
                    #对信号形状进行处理：如果信号形状和解剖或信号名称同时出现
                    elif "xhxz" in flags and ("jp" in flags or "xhmc" in flags):
                        WriteToCell(xhxz_cell,words[flags.index("xhxz")])
                    #对信号进行提取
                except:
                    pass
                
                #对信号强度进行提取
                if "xhqd" in flags:
                    if flags.count("xhmc")==1:
                        
                        if "T1" in words or "T1WI" in words:
                            to_cell = T1_cell
                        elif "T2" in words or "T2WI" in words:
                            to_cell = T2_cell
                        elif "DWI" in words:
                            to_cell = DWI_cell
                        elif "SWI" in words:
                            to_cell = SWI_cell
                        elif "FLAIR" in words:
                            to_cell = Flair_cell
                        elif "ADC" in words:
                            to_cell = ADC_cell
                        else:
                            continue
                        if "czxf" in flags:
                            WriteToCell(to_cell,"未见"+ words[flags.index("xhqd")]+"信号")
                        else:
                            WriteToCell(to_cell,words[flags.index("xhqd")]+"信号")
                
        print("\n")
    #保存
    book.save(path+out_filename)
    print("\nfinished")
main()