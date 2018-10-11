# -*- coding: utf-8 -*-
"""
Created on Thu Oct 11 10:33:27 2018

提取诊断中解剖部位
@author: 朱诚锐
"""

import jieba
import jieba.posseg as pseg
from openpyxl import load_workbook
import re
#载入自定义词典
jieba.load_userdict("User_Dict.txt")
#文件路径
path = r"C:\Users\朱诚锐\OneDrive\DEEPWISE\问诊平台\数据\征象提取"
#文件名
filename = r"\z-python处理后.xlsx"
dirc = path+filename

book = load_workbook(dirc)
sheet = book["Sheet1"]

content_col = 5
target_col = 9
'''
for row in range(2,sheet.max_row+1):
    diagnosis = sheet.cell(row=row,column=content_col).value.split("\n")
    position = []
    position_cell =  sheet.cell(row=row,column=target_col)
    for each_dig in diagnosis:
        if "海绵" in each_dig:
            result = pseg.cut(each_dig,HMM=False)
            for w in result:
                if w.flag == "jp":
                    if w.word not in position:
                        position.append(w.word)
    
                position_cell.value = "、".join(position)
'''
for row in range(2,sheet.max_row+1):
    diagnosis = sheet.cell(row=row,column=content_col).value.split("\n")
    position = []
    position_cell =  sheet.cell(row=row,column=target_col)
    for each_dig in diagnosis:
        if "海绵" in each_dig:
            if "多发" in each_dig:
                position_cell.value = "多发"


book.save(path+"\z-python处理后.xlsx")